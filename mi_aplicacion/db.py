import pg8000
from django.conf import settings
import base64
import pandas as pd
from django.http import HttpResponse 
from io import BytesIO
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment
from PIL import Image as PILImage


db_config = {
    'database': settings.DATABASES['default']['NAME'],
    'user': settings.DATABASES['default']['USER'],
    'password': settings.DATABASES['default']['PASSWORD'],
    'host': settings.DATABASES['default']['HOST'],
    'port': settings.DATABASES['default']['PORT']
}

def conectar():
    try:
        conexion = pg8000.connect(**db_config)
        return conexion
    except Exception as e:
        print(f"Error al conectar a la base de datos: {e}")
        return None

def validar_usuario(usuario, contrasena):
    conexion = conectar()
    if not conexion:
        return False

    try:
        cursor = conexion.cursor()
        query = "SELECT 1 FROM usuarios WHERE usuario = %s AND contrasena = %s"
        cursor.execute(query, (usuario, contrasena))
        result = cursor.fetchone()
        cursor.close()
        conexion.close()
        return result is not None
    except Exception as e:
        print(f"Error al validar usuario: {e}")
        return False

def obtener_gerencias():
    conexion = conectar()
    if not conexion:
        return []

    try:
        cursor = conexion.cursor()
        query = "SELECT id, nombre FROM Gerencia"  
        cursor.execute(query)
        gerencias = cursor.fetchall()
        cursor.close()
        conexion.close()
        return gerencias
    except Exception as e:
        print(f"Error al obtener gerencias: {e}")
        return []

def almacenar_asignacion(codigo, fecha_asignacion, gerencia_id, qr_code):
    conexion = conectar()
    if not conexion:
        return False

    try:
        cursor = conexion.cursor()
        query = """
        INSERT INTO Asignaciones (codigo, fecha_asignacion, gerencia, qr_code)
        VALUES (%s, %s, %s, %s)
        """
        cursor.execute(query, (codigo, fecha_asignacion, gerencia_id, qr_code))
        conexion.commit()
        cursor.close()
        conexion.close()
        return True
    except Exception as e:
        print(f"Error al almacenar la asignación: {e}")
        return False

def obtener_productos():
    conexion = conectar()
    if not conexion:
        return []

    try:
        cursor = conexion.cursor()
        query = """
        SELECT r.id_registro, r.producto, m.nombre AS marca, r.descripcion, r.color, g.nombre AS gerencia, r.fecha_registro
        FROM Registro r
        JOIN Marca m ON r.marca = m.id
        JOIN Gerencia g ON r.gerencia = g.id
        """
        cursor.execute(query)
        productos = cursor.fetchall()
        cursor.close()
        conexion.close()
        return productos
    except Exception as e:
        print(f"Error al obtener productos: {e}")
        return []
    
def Consulta_Bienes_Asignados():
    conexion = conectar()
    if not conexion:
        return []

    try:
        cursor = conexion.cursor()
        query = """
        SELECT a.id_asignacion, a.codigo, a.fecha_Asignacion, g.nombre AS gerencia, a.qr_code FROM asignaciones a 
        JOIN gerencia g ON a.gerencia = g.id
        """
        cursor.execute(query)
        asignaciones = cursor.fetchall()
        cursor.close()
        conexion.close()
        asignaciones_convertidas = [] 
        for asignacion in asignaciones: 
            asignacion_convertida = list(asignacion) 
            if asignacion_convertida[4]: 
                asignacion_convertida[4] = base64.b64encode(asignacion_convertida[4]).decode('utf-8') 
                asignaciones_convertidas.append(asignacion_convertida) 
        return asignaciones_convertidas
    except Exception as e:
        print(f"Error al obtener asiganciones: {e}")
        return []

def consulta_disponibles():
    conexion = conectar()
    if not conexion:
        return []

    try:
        cursor = conexion.cursor()
        query = """SELECT r.Id_Registro, m.Nombre AS Marca, r.Producto, r.Color, r.Modelo, r.Fecha_Registro, 
        c.categoria AS Categoria, r.Foto, r.Cantidad FROM Registro r 
        JOIN Marca m ON r.Marca = m.Id
        JOIN Categoria c ON r.Categoria = c.Id;"""  
        cursor.execute(query)
        disponibles = cursor.fetchall()
        cursor.close()
        conexion.close()
        registros_convertidos = [] 
        for registros in disponibles: 
            registro_convertido = list(registros) 
            if registro_convertido[7]: 
                registro_convertido[7] = base64.b64encode(registro_convertido[7]).decode('utf-8') 
                registros_convertidos.append(registro_convertido) 
        return registros_convertidos
    except Exception as e:
        print(f"Error al obtener Bienes Disponibles: {e}")
        return []

def obtener_disponibles():
    conexion = conectar()
    if not conexion:
        return []

    try:
        cursor = conexion.cursor()
        query = """SELECT r.Id_Registro, (r.Producto || ' ' || m.Nombre || ' ' || r.Color|| ' (' || r.Modelo|| ')') AS Producto, r.Fecha_Registro, 
        c.categoria AS Categoria, r.Foto, r.Cantidad FROM Registro r 
        JOIN Marca m ON r.Marca = m.Id
        JOIN Categoria c ON r.Categoria = c.Id;"""  
        cursor.execute(query)
        bien_disponible = cursor.fetchall()
        cursor.close()
        conexion.close()
        disponibles_convertidos = [] 
        for disponible in bien_disponible: 
            disponible_convertido = list(disponible) 
            if disponible_convertido[4]:
                disponible_convertido[4]= base64.b64encode(disponible_convertido[4]).decode('utf-8') 
                disponibles_convertidos.append(disponible_convertido) 
        return  disponibles_convertidos
    except Exception as e:
        print(f"Error al obtener Bienes Disponibles: {e}")
        return []


def consulta_desincorporacion():
    conexion = conectar()
    if not conexion:
        return []

    try:
        cursor = conexion.cursor()
        query = """SELECT 
        d.Id_Des, 
        a.codigo AS codigo_bien, 
        g.nombre AS Gerencia, 
        a.fecha_Asignacion,
        d.motivo_retiro,
        d.fecha_retiro,
        a.qr_code, 
        (al.estado || ', ' || al.municipio || ', ' || al.parroquia || ', ' || al.lugar) AS Ubicacion_Final 
        FROM 
        Desincorporacion d 
        JOIN 
        Asignaciones a ON d.Bien_Retirar = a.id_asignacion 
        JOIN 
        Gerencia g ON a.gerencia = g.id
        JOIN 
        Almacen al ON CAST(d.ubicación_final AS INTEGER) = al.id;"""  
        cursor.execute(query)
        desincorporacion= cursor.fetchall()
        cursor.close()
        conexion.close()
        return desincorporacion
    except Exception as e:
        print(f"Error al obtener Bienes desincorporados: {e}")
        return []

def obtener_categorias():
    conexion = conectar()
    if not conexion:
        return []

    try:
        cursor = conexion.cursor()
        query = "SELECT id, categoria FROM categoria"  
        cursor.execute(query)
        categorias = cursor.fetchall()
        cursor.close()
        conexion.close()
        return categorias
    except Exception as e:
        print(f"Error al obtener categorias: {e}")
        return []
    
def obtener_marcas():
    conexion = conectar()
    if not conexion:
        return []

    try:
        cursor = conexion.cursor()
        query = "SELECT id, nombre FROM marca"  
        cursor.execute(query)
        marcas = cursor.fetchall()
        cursor.close()
        conexion.close()
        return marcas
    except Exception as e:
        print(f"Error al obtener marcas: {e}")
        return []

def almacenar_registro(marca, producto, color, serial, fecha_registro, categoria, foto, cantidad):
    conexion = conectar()
    if not conexion:
        return False

    try:
        cursor = conexion.cursor()
        query = """
        INSERT INTO Registro (Marca, Producto, Color, Modelo, Fecha_Registro, Categoria, Foto, Cantidad)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """
        cursor.execute(query, (marca, producto, color, serial, fecha_registro, categoria, foto, cantidad))
        conexion.commit()
        cursor.close()
        conexion.close()
        return True
    except Exception as e:
        print(f"Error al almacenar el registro: {e}")
        return False
    
def obtener_Bienes():
    conexion = conectar()
    if not conexion:
        return []

    try:
        cursor = conexion.cursor()
        query = """SELECT id_registro, marca, producto, color, Modelo, fecha_registro, categoria, 
        foto, cantidad FROM registro"""  
        cursor.execute(query)
        bienes = cursor.fetchall()
        cursor.close()
        conexion.close()
        return bienes
    except Exception as e:
        print(f"Error al obtener marcas: {e}")
        return []

def guardar_Bienes_Asignados(request):
    conexion = conectar()
    if not conexion:
        return HttpResponse("No se pudo conectar a la base de datos", status=500)

    try:
        cursor = conexion.cursor()
        query = """
        SELECT a.id, a.codigo, a.fecha_Asignacion, g.nombre AS gerencia, a.qr_code 
        FROM asignacion a 
        JOIN gerencia g ON a.gerencia = g.id
        """
        cursor.execute(query)
        asignaciones = cursor.fetchall()
        cursor.close()
        conexion.close()
        
        # Crear libro de trabajo y hoja
        wb = Workbook()
        ws = wb.active
        ws.title = "Bienes Asignados"
        
        # Escribir encabezados
        encabezados = ["id", "codigo", "fecha_Asignacion", "gerencia", "qr_code"]
        ws.append(encabezados)
        for cell in ws[1]:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Escribir datos y manejar imágenes
        for asignacion in asignaciones:
            fila = list(asignacion[:4])  # Datos sin la imagen
            ws.append(fila)
            for cell in ws[ws.max_row]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            qr_code = asignacion[4]
            if qr_code:
                try:
                    img_file = BytesIO(qr_code)  # Usar directamente el contenido binario
                    # Usar PIL para abrir y redimensionar la imagen
                    pil_image = PILImage.open(img_file)
                    pil_image = pil_image.resize((80, 80), PILImage.Resampling.LANCZOS)
                    img_file = BytesIO()
                    pil_image.save(img_file, format="PNG")  # Guardar como PNG en un nuevo BytesIO
                    img_file.seek(0)  # Reinicia el puntero del archivo
                    img = Image(img_file)
                    img.anchor = f"E{ws.max_row}"
                    ws.add_image(img)
                    # Ajustar la altura de la fila
                    ws.row_dimensions[ws.max_row].height = 60  # Ajusta esto según sea necesario
                except Exception as e:
                    print(f"Error al procesar la imagen para la fila {asignacion[0]}: {e}")
        
        # Autoajustar el ancho de las columnas
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length + 2  # Añadir margen
        
        # Guardar el archivo Excel en memoria
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Crear una respuesta HTTP con el archivo Excel
        response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=bienes_asignados.xlsx'
        
        return response
        
    except Exception as e:
        return HttpResponse(f"Error al exportar datos: {e}", status=500)

def consulta_registro_por_codigo(codigo):
    conexion = conectar()
    if not conexion:
        return None

    try:
        cursor = conexion.cursor()
        query = """SELECT m.Nombre AS Marca, r.Producto, r.Color, r.Modelo, r.Foto 
                   FROM Registro r 
                   JOIN Marca m ON r.Marca = m.Id
                   WHERE r.id_registro = %s;"""
        cursor.execute(query, (codigo,))
        registro = cursor.fetchone()
        cursor.close()
        conexion.close()

        if registro:
            registro_convertido = list(registro)
            if registro_convertido[4]:  # Si hay foto
                registro_convertido[4] = base64.b64encode(registro_convertido[4]).decode('utf-8')
            return registro_convertido
        else:
            return None
    except Exception as e:
        print(f"Error al obtener registro por código: {e}")
        return None
